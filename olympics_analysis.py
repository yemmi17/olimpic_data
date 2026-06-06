from __future__ import annotations

import argparse
import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, PieChart, ScatterChart, Series, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


DEFAULT_CSV = Path("data") / "athlete_events.csv"
OUTPUT_DIR = Path("outputs")
OUTPUT_XLSX = OUTPUT_DIR / "olympics_report.xlsx"
SITE_DATA_JS = Path("site") / "data.js"

COUNTRY_GROUPS = {
    "Russia/USSR": {"RUS", "URS"},
    "USA": {"USA"},
}


def load_data(csv_path: Path) -> pd.DataFrame:
    df = pd.read_csv(csv_path)
    df["Medal"] = df["Medal"].replace({"NA": pd.NA})
    return df


def medal_events(df: pd.DataFrame) -> pd.DataFrame:
    medal_df = df[df["Medal"].notna()].copy()
    # One medal in a team event appears once per athlete in the raw dataset.
    return medal_df.drop_duplicates(["Year", "Season", "Event", "NOC", "Medal"])


def write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> tuple[int, int]:
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for col_idx, value in enumerate(row, start_col):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == start_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.alignment = Alignment(horizontal="center")
    autosize(ws)
    return start_row + len(df), start_col + len(df.columns) - 1


def autosize(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column].width = min(max_len + 2, 45)


def add_bar_chart(ws, title: str, min_col: int, min_row: int, max_col: int, max_row: int, anchor: str) -> None:
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = ws.cell(min_row, min_col).value
    chart.x_axis.title = "Количество"
    data = Reference(ws, min_col=min_col + 1, max_col=max_col, min_row=min_row, max_row=max_row)
    cats = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 22
    ws.add_chart(chart, anchor)


def add_column_chart(ws, title: str, min_col: int, min_row: int, max_col: int, max_row: int, anchor: str) -> None:
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "Количество"
    data = Reference(ws, min_col=min_col + 1, max_col=max_col, min_row=min_row, max_row=max_row)
    cats = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 22
    ws.add_chart(chart, anchor)


def add_line_chart(ws, title: str, min_col: int, min_row: int, max_col: int, max_row: int, anchor: str) -> None:
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Количество"
    chart.x_axis.title = "Год"
    data = Reference(ws, min_col=min_col + 1, max_col=max_col, min_row=min_row, max_row=max_row)
    cats = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 22
    ws.add_chart(chart, anchor)


def add_pie_chart(ws, title: str, min_col: int, min_row: int, max_row: int, anchor: str, doughnut: bool = False) -> None:
    chart = DoughnutChart() if doughnut else PieChart()
    chart.title = title
    data = Reference(ws, min_col=min_col + 1, min_row=min_row, max_row=max_row)
    cats = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.height = 10
    chart.width = 14
    ws.add_chart(chart, anchor)


def add_scatter_chart(ws, title: str, x_col: int, y_col: int, min_row: int, max_row: int, anchor: str) -> None:
    chart = ScatterChart()
    chart.title = title
    chart.x_axis.title = ws.cell(1, x_col).value
    chart.y_axis.title = ws.cell(1, y_col).value
    xvalues = Reference(ws, min_col=x_col, min_row=min_row, max_row=max_row)
    yvalues = Reference(ws, min_col=y_col, min_row=min_row, max_row=max_row)
    series = Series(yvalues, xvalues, title_from_data=False)
    chart.series.append(series)
    chart.height = 12
    chart.width = 18
    ws.add_chart(chart, anchor)


def records(df: pd.DataFrame) -> list[dict]:
    clean = df.replace({pd.NA: None}).where(pd.notna(df), None)
    return clean.to_dict(orient="records")


def write_site_data(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    js = "window.OLYMPICS_DATA = "
    js += json.dumps(payload, ensure_ascii=False, allow_nan=False, indent=2)
    js += ";\n"
    path.write_text(js, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Olympic Games dataset analysis.")
    parser.add_argument("--csv", type=Path, default=DEFAULT_CSV, help="Path to athlete_events.csv")
    parser.add_argument("--output", type=Path, default=OUTPUT_XLSX, help="Output xlsx report path")
    parser.add_argument("--site-data", type=Path, default=SITE_DATA_JS, help="Output JS data file for the site")
    args = parser.parse_args()

    df = load_data(args.csv)
    event_medals = medal_events(df)
    args.output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    games = (
        df[["Games", "Year", "Season", "City"]]
        .drop_duplicates()
        .sort_values(["Year", "Season"])
        .reset_index(drop=True)
    )
    games_count = games.groupby("Season", as_index=False).agg(Count=("Games", "count"))

    first_curling_year = int(df.loc[(df["Sport"] == "Curling") & (df["Medal"] == "Gold"), "Year"].min())
    first_curling_gold = (
        df[(df["Sport"] == "Curling") & (df["Medal"] == "Gold") & (df["Year"] == first_curling_year)]
        [["Year", "Games", "Team", "NOC", "Name", "Event"]]
        .drop_duplicates()
        .sort_values(["Event", "Name"])
        .reset_index(drop=True)
    )

    russia_sochi_gold_athletes = (
        df[(df["Year"] == 2014) & (df["City"] == "Sochi") & (df["NOC"] == "RUS") & (df["Medal"] == "Gold")]
        [["Name", "Sex", "Age", "Sport", "Event"]]
        .drop_duplicates()
        .sort_values(["Sport", "Event", "Name"])
        .reset_index(drop=True)
    )
    russia_sochi_gold_unique = (
        russia_sochi_gold_athletes.groupby(["Name", "Sex"], as_index=False)
        .agg(
            Age=("Age", "min"),
            Gold_events=("Event", "count"),
            Sports=("Sport", lambda values: ", ".join(sorted(set(values)))),
            Events=("Event", lambda values: "; ".join(sorted(set(values)))),
        )
        .sort_values(["Gold_events", "Name"], ascending=[False, True])
        .reset_index(drop=True)
    )

    comparison_rows = []
    for year, city in [(2014, "Sochi"), (1980, "Moskva")]:
        subset = event_medals[(event_medals["Year"] == year) & (event_medals["City"] == city)]
        total = len(subset)
        for label, nocs in COUNTRY_GROUPS.items():
            medals = len(subset[subset["NOC"].isin(nocs)])
            comparison_rows.append(
                {
                    "Year": year,
                    "City": city,
                    "Country group": label,
                    "Medal events": medals,
                    "Share of all medal events, %": round(medals / total * 100, 2) if total else 0,
                }
            )
    country_comparison = pd.DataFrame(comparison_rows)

    age_distribution = (
        df[df["Age"].notna()]
        .assign(Age=lambda x: x["Age"].astype(int))
        .groupby("Age", as_index=False)
        .agg(Athlete_rows=("ID", "count"), Unique_athletes=("ID", "nunique"))
        .sort_values("Age")
    )
    gender_by_age = (
        df[df["Age"].notna()]
        .assign(Age=lambda x: x["Age"].astype(int))
        .groupby(["Sex", "Age"], as_index=False)
        .agg(Athlete_rows=("ID", "count"))
        .sort_values(["Sex", "Age"])
    )
    gender_total = df.groupby("Sex", as_index=False).agg(Athlete_rows=("ID", "count"))
    female_line = (
        df[df["Sex"] == "F"]
        .groupby("Year", as_index=False)
        .agg(Female_athlete_rows=("ID", "count"), Unique_female_athletes=("ID", "nunique"))
        .sort_values("Year")
    )
    season_participants = df.groupby("Season", as_index=False).agg(Athlete_rows=("ID", "count"))

    top_countries = (
        event_medals.groupby(["NOC", "Medal"], as_index=False)
        .size()
        .pivot(index="NOC", columns="Medal", values="size")
        .fillna(0)
        .astype(int)
        .reset_index()
    )
    for medal in ["Gold", "Silver", "Bronze"]:
        if medal not in top_countries:
            top_countries[medal] = 0
    top_countries["Total"] = top_countries[["Gold", "Silver", "Bronze"]].sum(axis=1)
    top_countries = top_countries[["NOC", "Gold", "Silver", "Bronze", "Total"]].sort_values(
        ["Total", "Gold", "Silver"], ascending=False
    ).head(20)

    rus_summer_sports = (
        event_medals[(event_medals["NOC"] == "RUS") & (event_medals["Season"] == "Summer")]
        .groupby("Sport", as_index=False)
        .size()
        .rename(columns={"size": "Medal events"})
        .sort_values("Medal events", ascending=False)
        .head(10)
    )

    top_athletes = (
        df[df["Medal"].notna()]
        .drop_duplicates(["ID", "Year", "Season", "Event", "Medal"])
        .groupby(["ID", "Name", "Team", "NOC"], as_index=False)
        .agg(Medals=("Medal", "count"), Gold=("Medal", lambda s: (s == "Gold").sum()))
        .sort_values(["Medals", "Gold", "Name"], ascending=[False, False, True])
        .head(10)
        .drop(columns=["ID"])
    )

    scatter_df = df[["Name", "Sport", "Age", "Height", "Weight"]].dropna().drop_duplicates().reset_index(drop=True)
    extremes = pd.DataFrame(
        [
            {"Metric": "Max height", **scatter_df.loc[scatter_df["Height"].idxmax()].to_dict()},
            {"Metric": "Max weight", **scatter_df.loc[scatter_df["Weight"].idxmax()].to_dict()},
            {"Metric": "Max age", **scatter_df.loc[scatter_df["Age"].idxmax()].to_dict()},
        ]
    )

    older_champions = (
        df[(df["Medal"] == "Gold") & (df["Age"] > 60)]
        [["Name", "Sex", "Age", "Team", "NOC", "Year", "Season", "Sport", "Event"]]
        .drop_duplicates()
        .sort_values(["Age", "Year"], ascending=[False, True])
        .reset_index(drop=True)
    )
    older_sports = (
        df[(df["Medal"].notna()) & (df["Age"] > 60)]
        .drop_duplicates(["ID", "Year", "Season", "Event", "Medal"])
        .groupby("Sport", as_index=False)
        .size()
        .rename(columns={"size": "Medals by athletes over 60"})
        .sort_values("Medals by athletes over 60", ascending=False)
    )

    ws = wb.create_sheet("Summary")
    summary = pd.DataFrame(
        [
            {"Question": "Winter Olympic Games through 2016", "Answer": int(games_count.loc[games_count["Season"] == "Winter", "Count"].iloc[0])},
            {"Question": "Summer Olympic Games through 2016", "Answer": int(games_count.loc[games_count["Season"] == "Summer", "Count"].iloc[0])},
            {"Question": "First year with curling gold in dataset", "Answer": first_curling_year},
            {"Question": "Unique Russian gold medalist athletes in Sochi 2014", "Answer": len(russia_sochi_gold_unique)},
            {"Question": "Russian gold athlete-event rows in Sochi 2014", "Answer": len(russia_sochi_gold_athletes)},
        ]
    )
    write_df(ws, summary)
    write_df(ws, games_count, start_row=8)

    sheets = {
        "Games": games,
        "First curling gold": first_curling_gold,
        "Russia Sochi unique gold": russia_sochi_gold_unique,
        "Russia Sochi gold": russia_sochi_gold_athletes,
        "Russia USA comparison": country_comparison,
        "Age distribution": age_distribution,
        "Gender by age": gender_by_age,
        "Gender total": gender_total,
        "Female by year": female_line,
        "Season participants": season_participants,
        "Top countries medals": top_countries,
        "RUS summer sports": rus_summer_sports,
        "Top athletes": top_athletes,
        "Scatter source": scatter_df.head(5000),
        "Extremes": extremes,
        "Gold champions over 60": older_champions,
        "Sports over 60 medals": older_sports,
    }

    site_payload = {
        "summary": records(summary),
        "gamesCount": records(games_count),
        "firstCurlingGold": records(first_curling_gold),
        "russiaSochiGoldUnique": records(russia_sochi_gold_unique),
        "countryComparison": records(country_comparison),
        "ageDistribution": records(age_distribution),
        "genderByAge": records(gender_by_age),
        "genderTotal": records(gender_total),
        "femaleLine": records(female_line),
        "seasonParticipants": records(season_participants),
        "topCountries": records(top_countries.head(12)),
        "rusSummerSports": records(rus_summer_sports),
        "topAthletes": records(top_athletes),
        "scatter": records(scatter_df.sample(n=min(1800, len(scatter_df)), random_state=7)),
        "extremes": records(extremes),
        "olderChampions": records(older_champions),
        "olderSports": records(older_sports),
    }
    write_site_data(args.site_data, site_payload)

    for name, table in sheets.items():
        ws = wb.create_sheet(name[:31])
        write_df(ws, table)

    add_column_chart(wb["Age distribution"], "Распределение спортсменов по возрасту", 1, 1, 3, len(age_distribution) + 1, "F2")
    add_pie_chart(wb["Gender total"], "Гендерное распределение", 1, 1, len(gender_total) + 1, "D2")
    add_line_chart(wb["Female by year"], "Динамика участия женщин", 1, 1, 3, len(female_line) + 1, "F2")
    add_pie_chart(wb["Season participants"], "Участники летних и зимних игр", 1, 1, len(season_participants) + 1, "D2", doughnut=True)
    add_bar_chart(wb["Top countries medals"], "Страны с наибольшим количеством медалей", 1, 1, 5, len(top_countries) + 1, "G2")
    add_pie_chart(wb["RUS summer sports"], "Топ-10 летних видов спорта РФ по наградам", 1, 1, len(rus_summer_sports) + 1, "D2", doughnut=True)
    add_scatter_chart(wb["Scatter source"], "Рост и вес", 4, 5, 2, len(wb["Scatter source"]["A"]), "G2")
    add_scatter_chart(wb["Scatter source"], "Возраст и вес", 3, 5, 2, len(wb["Scatter source"]["A"]), "G25")
    add_scatter_chart(wb["Scatter source"], "Возраст и рост", 3, 4, 2, len(wb["Scatter source"]["A"]), "G48")
    add_column_chart(wb["Gold champions over 60"], "Олимпийские чемпионы старше 60 лет", 3, 1, 3, len(older_champions) + 1, "K2")
    add_bar_chart(wb["Sports over 60 medals"], "Виды спорта: медали спортсменов старше 60 лет", 1, 1, 2, len(older_sports) + 1, "D2")

    wb.save(args.output)
    print(f"Report saved: {args.output.resolve()}")
    print(f"Site data saved: {args.site_data.resolve()}")
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
